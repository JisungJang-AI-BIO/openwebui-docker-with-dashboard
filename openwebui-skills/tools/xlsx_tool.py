"""
title: XLSX Spreadsheet Tool
author: Internal Team
author_url: https://github.com/anthropics/skills
description: Read, create, edit, and convert Excel spreadsheets (.xlsx).
    Use when user uploads a spreadsheet or requests Excel operations such as
    reading data, creating workbooks with formulas, editing cells, or converting formats.
required_open_webui_version: 0.8.5
requirements: openpyxl, pandas, xlsxwriter
version: 1.0.0
licence: MIT
"""

import asyncio
import base64
import io
import json
import os
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any, Callable, Optional

from pydantic import BaseModel, Field


# =============================================================================
# EventEmitter Helper
# =============================================================================
class EventEmitter:
    def __init__(self, event_emitter: Callable[[dict], Any] = None, user: dict = None):
        self.event_emitter = event_emitter
        self.user = user or {}

    async def emit(self, description="Unknown State", status="in_progress", done=False):
        if self.event_emitter:
            await self.event_emitter(
                {"type": "status", "data": {"status": status, "description": description, "done": done}}
            )

    async def progress(self, description: str):
        await self.emit(description)

    async def success(self, description: str):
        await self.emit(description, "success", True)

    async def error(self, description: str):
        await self.emit(description, "error", True)

    async def send_file_link(self, file_path: str, filename: str, mime_type: str = None):
        """Upload file to OpenWebUI Files API and emit download link.
        Falls back to base64 data URI if the API is unavailable.
        """
        if not self.event_emitter or not os.path.exists(file_path):
            return
        if mime_type is None:
            ext = Path(filename).suffix.lower()
            mime_map = {
                ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
                ".pdf": "application/pdf", ".html": "text/html", ".txt": "text/plain",
                ".csv": "text/csv", ".odt": "application/vnd.oasis.opendocument.text",
                ".rtf": "application/rtf", ".png": "image/png",
                ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".gif": "image/gif",
            }
            mime_type = mime_map.get(ext, "application/octet-stream")

        is_image = mime_type.startswith("image/")
        url = await self._upload_to_openwebui(file_path, filename, mime_type)

        if url:
            if is_image:
                content = f"\n\n📎 **{filename}**\n\n![{filename}]({url})\n\n[Download {filename}]({url})\n"
            else:
                content = f"\n\n📎 **{filename}**\n\n[Download {filename}]({url})\n"
        else:
            with open(file_path, "rb") as f:
                b64 = base64.b64encode(f.read()).decode("utf-8")
            data_uri = f"data:{mime_type};base64,{b64}"
            if is_image:
                content = f"\n\n📎 **{filename}**\n\n![{filename}]({data_uri})\n\n[Download {filename}]({data_uri})\n"
            else:
                content = f"\n\n📎 **{filename}**\n\n[Download {filename}]({data_uri})\n"

        await self.event_emitter({"type": "message", "data": {"content": content}})

    async def _upload_to_openwebui(self, file_path: str, filename: str, mime_type: str) -> Optional[str]:
        """Upload file via OpenWebUI's internal Files API. Returns download URL or None."""
        try:
            import httpx
            import jwt

            user_id = self.user.get("id", "")
            secret = os.environ.get("WEBUI_SECRET_KEY", "")
            if not user_id or not secret:
                return None

            token = jwt.encode({"id": user_id}, secret, algorithm="HS256")
            if isinstance(token, bytes):
                token = token.decode("utf-8")

            port = os.environ.get("PORT", "8080")
            with open(file_path, "rb") as f:
                resp = httpx.post(
                    f"http://localhost:{port}/api/v1/files/",
                    files={"file": (filename, f, mime_type)},
                    headers={"Authorization": f"Bearer {token}"},
                    timeout=30.0,
                )

            if resp.status_code in (200, 201):
                file_id = resp.json().get("id")
                if file_id:
                    return f"/api/v1/files/{file_id}/content"
            return None
        except Exception:
            return None


# =============================================================================
# Main Tool Class
# =============================================================================
class Tools:
    class Valves(BaseModel):
        TEMP_DIR: str = Field(default="/tmp/openwebui-xlsx", description="Temporary file storage path")
        MAX_FILE_SIZE_MB: int = Field(default=50, description="Maximum file size in MB")
        SCRIPTS_DIR: str = Field(
            default="",
            description="Absolute path to Anthropic xlsx scripts (e.g. /app/OpenWebUI-Skills/vendor/xlsx). Leave empty for fallback mode.",
        )
        LIBREOFFICE_PATH: str = Field(default="soffice", description="LibreOffice binary path")

    def __init__(self):
        self.valves = self.Valves()
        self.citation = False

    # =========================================================================
    # Tool Methods
    # =========================================================================

    async def read_xlsx(
        self,
        mode: str = "text",
        sheet_name: str = "",
        max_rows: int = 100,
        __files__: list = None,
        __user__: dict = None,
        __event_emitter__: Callable[[dict], Any] = None,
    ) -> str:
        """
        Read and extract content from an uploaded Excel file (.xlsx).

        Modes:
        - "text": Data as markdown tables (values only)
        - "formulas": Show formulas instead of computed values
        - "metadata": Sheet names, dimensions, named ranges

        :param mode: "text", "formulas", or "metadata"
        :param sheet_name: Specific sheet name to read (empty = all sheets)
        :param max_rows: Maximum rows to read per sheet (default 100)
        :return: Extracted spreadsheet content
        """
        emitter = EventEmitter(__event_emitter__, __user__)
        await emitter.progress("Reading XLSX...")

        if not __files__:
            await emitter.error("No file uploaded")
            return "Error: No file uploaded. Please upload an .xlsx file."

        file_path = self._resolve_xlsx_file(__files__)
        if not file_path:
            await emitter.error("No XLSX file found")
            return "Error: No .xlsx file found."

        try:
            import openpyxl

            if mode == "metadata":
                return self._read_metadata(file_path)

            data_only = mode != "formulas"
            wb = openpyxl.load_workbook(file_path, data_only=data_only)
            sheets = [sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames

            parts = [f"## Workbook: {Path(file_path).name}\n"]
            parts.append(f"**Sheets**: {', '.join(wb.sheetnames)}\n")

            for sn in sheets:
                ws = wb[sn]
                parts.append(f"### Sheet: {sn} ({ws.max_row}x{ws.max_column})")

                if ws.max_row == 0 or ws.max_column == 0:
                    parts.append("*(empty sheet)*\n")
                    continue

                rows_to_read = min(ws.max_row, max_rows)
                for r in range(1, rows_to_read + 1):
                    cells = []
                    for c in range(1, ws.max_column + 1):
                        cell = ws.cell(row=r, column=c)
                        val = cell.value
                        if val is None:
                            cells.append("")
                        elif isinstance(val, (int, float)):
                            cells.append(str(val))
                        else:
                            cells.append(str(val))
                    parts.append("| " + " | ".join(cells) + " |")
                    if r == 1:
                        parts.append("| " + " | ".join(["---"] * len(cells)) + " |")

                if ws.max_row > max_rows:
                    parts.append(f"\n*(showing {max_rows} of {ws.max_row} rows)*")
                parts.append("")

            wb.close()
            await emitter.success(f"Read {len(sheets)} sheet(s)")
            return "\n".join(parts)

        except Exception as e:
            await emitter.error(f"Read error: {str(e)}")
            return f"Error reading XLSX: {str(e)}"

    async def create_xlsx(
        self,
        content: str,
        filename: str = "spreadsheet.xlsx",
        __user__: dict = None,
        __event_emitter__: Callable[[dict], Any] = None,
    ) -> str:
        """
        Create a new Excel spreadsheet.

        Content format — use sheet headers and table notation:
        ```
        ## Sheet: Revenue
        | Year | Q1 | Q2 | Q3 | Q4 | Total |
        | 2024 | 100 | 120 | 130 | 150 | =SUM(B2:E2) |
        | 2025 | 110 | 135 | 145 | 170 | =SUM(B3:E3) |

        ## Sheet: Summary
        | Metric | Value |
        | Total Revenue | =Revenue!F2+Revenue!F3 |
        ```

        CRITICAL RULES for formulas:
        - ALWAYS use Excel formulas (=SUM, =AVERAGE, etc.) — NEVER hardcode calculated values
        - Formula cells should be black text; hardcoded input cells blue text
        - Number formats: currency as $#,##0, percentages as 0.0%, zeros displayed as "-"
        - Cross-sheet references: =SheetName!A1

        :param content: Spreadsheet content specification with formulas
        :param filename: Output filename
        :return: Status with download link
        """
        emitter = EventEmitter(__event_emitter__, __user__)
        await emitter.progress("Creating XLSX...")

        try:
            import openpyxl
            from openpyxl.styles import Font, numbers

            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            sheets_data = self._parse_sheet_content(content)

            for sheet_name, rows in sheets_data.items():
                ws = wb.create_sheet(title=sheet_name)
                for r_idx, row in enumerate(rows, 1):
                    for c_idx, val in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx)
                        val = val.strip()

                        if val.startswith("="):
                            # Formula
                            cell.value = val
                            cell.font = Font(color="000000")
                        elif self._is_number(val):
                            cell.value = float(val) if "." in val else int(val)
                            cell.font = Font(color="0000FF")  # Blue = hardcoded input
                        else:
                            cell.value = val

                        # Bold header row
                        if r_idx == 1:
                            cell.font = Font(bold=True)

                # Auto-fit column widths (approximate)
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

            if not wb.sheetnames:
                wb.create_sheet("Sheet1")

            temp_dir = self._ensure_temp_dir()
            output_path = os.path.join(temp_dir, filename)
            wb.save(output_path)
            wb.close()

            await emitter.send_file_link(output_path, filename)
            await emitter.success(f"Created {filename}")

            file_size = os.path.getsize(output_path)
            return f"Created '{filename}' ({file_size:,} bytes, {len(sheets_data)} sheet(s))."

        except Exception as e:
            await emitter.error(f"Creation error: {str(e)}")
            return f"Error creating XLSX: {str(e)}"

    async def edit_xlsx(
        self,
        operation: str,
        parameters: str = "",
        __files__: list = None,
        __user__: dict = None,
        __event_emitter__: Callable[[dict], Any] = None,
    ) -> str:
        """
        Edit an uploaded Excel file.

        Operations:
        - "set_cell": Set cell value. parameters = "SheetName!A1|||value" or "A1|||value" (first sheet)
        - "add_sheet": Add a new sheet. parameters = sheet name
        - "delete_sheet": Delete a sheet. parameters = sheet name
        - "add_row": Append a row. parameters = "SheetName|||val1,val2,val3" or "val1,val2,val3"

        :param operation: Operation name
        :param parameters: Operation-specific parameters
        :return: Edited file download link
        """
        emitter = EventEmitter(__event_emitter__, __user__)
        await emitter.progress(f"Editing XLSX: {operation}...")

        if not __files__:
            await emitter.error("No file uploaded")
            return "Error: No file uploaded."

        file_path = self._resolve_xlsx_file(__files__)
        if not file_path:
            return "Error: No .xlsx file found."

        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path)
            base_name = Path(file_path).stem

            if operation == "set_cell":
                parts = parameters.split("|||")
                if len(parts) != 2:
                    return "Error: format is 'SheetName!A1|||value' or 'A1|||value'"
                ref, val = parts
                if "!" in ref:
                    sn, cell_ref = ref.split("!", 1)
                    ws = wb[sn]
                else:
                    ws = wb.active
                    cell_ref = ref
                cell = ws[cell_ref]
                if val.startswith("="):
                    cell.value = val
                elif self._is_number(val):
                    cell.value = float(val) if "." in val else int(val)
                else:
                    cell.value = val
                msg = f"Set {ref} = {val}"

            elif operation == "add_sheet":
                wb.create_sheet(title=parameters.strip())
                msg = f"Added sheet '{parameters.strip()}'."

            elif operation == "delete_sheet":
                sn = parameters.strip()
                if sn in wb.sheetnames:
                    del wb[sn]
                    msg = f"Deleted sheet '{sn}'."
                else:
                    return f"Error: Sheet '{sn}' not found. Available: {', '.join(wb.sheetnames)}"

            elif operation == "add_row":
                parts = parameters.split("|||")
                if len(parts) == 2:
                    ws = wb[parts[0]]
                    values = parts[1].split(",")
                else:
                    ws = wb.active
                    values = parameters.split(",")
                row_data = []
                for v in values:
                    v = v.strip()
                    if v.startswith("="):
                        row_data.append(v)
                    elif self._is_number(v):
                        row_data.append(float(v) if "." in v else int(v))
                    else:
                        row_data.append(v)
                ws.append(row_data)
                msg = f"Added row with {len(row_data)} cells."

            else:
                return f"Error: Unknown operation '{operation}'. Available: set_cell, add_sheet, delete_sheet, add_row"

            temp_dir = self._ensure_temp_dir()
            output_filename = f"{base_name}_edited.xlsx"
            output_path = os.path.join(temp_dir, output_filename)
            wb.save(output_path)
            wb.close()

            await emitter.send_file_link(output_path, output_filename)
            await emitter.success(msg)
            return msg

        except Exception as e:
            await emitter.error(f"Edit error: {str(e)}")
            return f"Error: {str(e)}"

    async def recalculate(
        self,
        __files__: list = None,
        __user__: dict = None,
        __event_emitter__: Callable[[dict], Any] = None,
    ) -> str:
        """
        Recalculate all formulas in an Excel file using LibreOffice headless.
        This is needed because openpyxl does not compute formula results.

        :return: Recalculated file download link and any errors
        """
        emitter = EventEmitter(__event_emitter__, __user__)
        await emitter.progress("Recalculating formulas...")

        if not __files__:
            await emitter.error("No file uploaded")
            return "Error: No file uploaded."

        file_path = self._resolve_xlsx_file(__files__)
        if not file_path:
            return "Error: No .xlsx file found."

        try:
            temp_dir = self._ensure_temp_dir()
            scripts_dir = self.valves.SCRIPTS_DIR
            recalc_script = os.path.join(scripts_dir, "recalc.py") if scripts_dir else ""
            base_name = Path(file_path).stem

            if recalc_script and os.path.exists(recalc_script):
                # Use Anthropic's recalc.py
                output_path = os.path.join(temp_dir, f"{base_name}_recalc.xlsx")
                shutil.copy2(file_path, output_path)
                result = subprocess.run(
                    ["python", recalc_script, output_path],
                    capture_output=True, text=True, timeout=60,
                    cwd=scripts_dir,
                )
                output_text = result.stdout.strip()
                try:
                    report = json.loads(output_text)
                    msg = f"Recalculation: {report.get('status', 'done')}. Errors: {report.get('total_errors', 0)}"
                except json.JSONDecodeError:
                    msg = f"Recalculation done. {output_text}"
            else:
                # Fallback: LibreOffice macro
                output_path = os.path.join(temp_dir, f"{base_name}_recalc.xlsx")
                shutil.copy2(file_path, output_path)
                soffice = self.valves.LIBREOFFICE_PATH
                env = os.environ.copy()
                env["SAL_USE_VCLPLUGIN"] = "svp"
                try:
                    subprocess.run(
                        [soffice, "--headless", "--calc", "--convert-to", "xlsx", "--outdir", temp_dir, output_path],
                        capture_output=True, text=True, timeout=60, env=env,
                    )
                    msg = "Recalculated via LibreOffice (fallback mode)."
                except FileNotFoundError:
                    return f"Error: LibreOffice not found at '{soffice}'."

            if os.path.exists(output_path):
                await emitter.send_file_link(output_path, f"{base_name}_recalc.xlsx")
                await emitter.success("Recalculation complete")
                return msg
            return "Error: Recalculation failed."

        except Exception as e:
            await emitter.error(f"Recalc error: {str(e)}")
            return f"Error: {str(e)}"

    async def convert_xlsx(
        self,
        target_format: str = "csv",
        sheet_name: str = "",
        __files__: list = None,
        __user__: dict = None,
        __event_emitter__: Callable[[dict], Any] = None,
    ) -> str:
        """
        Convert an uploaded Excel file to another format.
        Supported: csv, pdf, html.

        :param target_format: "csv", "pdf", or "html"
        :param sheet_name: Specific sheet for CSV export (empty = active sheet)
        :return: Converted file or download link
        """
        emitter = EventEmitter(__event_emitter__, __user__)
        await emitter.progress(f"Converting to {target_format}...")

        if not __files__:
            await emitter.error("No file uploaded")
            return "Error: No file uploaded."

        file_path = self._resolve_xlsx_file(__files__)
        if not file_path:
            return "Error: No .xlsx file found."

        try:
            temp_dir = self._ensure_temp_dir()
            fmt = target_format.lower().strip(".")

            if fmt == "csv":
                return await self._convert_to_csv(file_path, temp_dir, sheet_name, emitter)
            elif fmt in ("pdf", "html"):
                return await self._convert_libreoffice(file_path, temp_dir, fmt, emitter)
            else:
                return f"Error: Unsupported format '{fmt}'. Supported: csv, pdf, html"

        except Exception as e:
            await emitter.error(f"Conversion error: {str(e)}")
            return f"Error: {str(e)}"

    # =========================================================================
    # Internal helpers
    # =========================================================================

    def _resolve_xlsx_file(self, files: list) -> Optional[str]:
        if not files:
            return None
        for f in files:
            info = f if isinstance(f, dict) else {"path": f}
            name = info.get("filename", info.get("name", str(info.get("path", ""))))
            if not name.lower().endswith((".xlsx", ".xls")):
                continue
            for key in ("path", "file_path", "url", "id"):
                path = info.get(key)
                if path and isinstance(path, str):
                    for c in [path, f"/app/backend/data/uploads/{path}", f"/app/backend/data/cache/{path}"]:
                        if os.path.exists(c):
                            return c
                    return path
        if files:
            f = files[0]
            if isinstance(f, dict):
                for key in ("path", "file_path"):
                    if key in f:
                        return f[key]
            elif isinstance(f, str):
                return f
        return None

    def _ensure_temp_dir(self) -> str:
        os.makedirs(self.valves.TEMP_DIR, exist_ok=True)
        return self.valves.TEMP_DIR

    def _is_number(self, s: str) -> bool:
        try:
            float(s)
            return True
        except (ValueError, TypeError):
            return False

    def _read_metadata(self, file_path: str) -> str:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True)
        parts = ["## Workbook Metadata\n"]
        parts.append(f"**File**: {Path(file_path).name}")
        parts.append(f"**Sheets**: {len(wb.sheetnames)}\n")

        for sn in wb.sheetnames:
            ws = wb[sn]
            parts.append(f"### {sn}")
            parts.append(f"  Dimensions: {ws.dimensions}")
            parts.append(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")

        wb.close()
        return "\n".join(parts)

    def _parse_sheet_content(self, content: str) -> dict:
        sheets = {}
        current_sheet = "Sheet1"
        current_rows = []

        for line in content.split("\n"):
            s = line.strip()
            if s.startswith("## Sheet:") or s.startswith("## sheet:"):
                if current_rows:
                    sheets[current_sheet] = current_rows
                current_sheet = s.split(":", 1)[1].strip()
                current_rows = []
            elif s.startswith("|") and s.endswith("|"):
                inner = s[1:-1].strip()
                if all(c in "-| :" for c in inner):
                    continue
                cells = [c.strip() for c in s.strip("|").split("|")]
                current_rows.append(cells)

        if current_rows:
            sheets[current_sheet] = current_rows

        return sheets

    async def _convert_to_csv(self, file_path, temp_dir, sheet_name, emitter) -> str:
        import pandas as pd

        xls = pd.ExcelFile(file_path)
        sn = sheet_name if sheet_name and sheet_name in xls.sheet_names else xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=sn)

        output_filename = f"{Path(file_path).stem}_{sn}.csv"
        output_path = os.path.join(temp_dir, output_filename)
        df.to_csv(output_path, index=False)

        await emitter.send_file_link(output_path, output_filename)
        await emitter.success(f"Converted to CSV")
        return f"Converted sheet '{sn}' to CSV ({os.path.getsize(output_path):,} bytes)."

    async def _convert_libreoffice(self, file_path, temp_dir, fmt, emitter) -> str:
        soffice = self.valves.LIBREOFFICE_PATH
        input_copy = os.path.join(temp_dir, Path(file_path).name)
        shutil.copy2(file_path, input_copy)
        env = os.environ.copy()
        env["SAL_USE_VCLPLUGIN"] = "svp"
        try:
            result = subprocess.run(
                [soffice, "--headless", "--convert-to", fmt, "--outdir", temp_dir, input_copy],
                capture_output=True, text=True, timeout=120, env=env,
            )
        except FileNotFoundError:
            return f"Error: LibreOffice not found at '{soffice}'."
        except subprocess.TimeoutExpired:
            return "Error: Conversion timed out."

        output_path = os.path.join(temp_dir, f"{Path(file_path).stem}.{fmt}")
        if os.path.exists(output_path):
            await emitter.send_file_link(output_path, Path(output_path).name)
            await emitter.success(f"Converted to {fmt}")
            return f"Converted to {fmt} ({os.path.getsize(output_path):,} bytes)."
        return f"Error: Conversion failed. {result.stderr}"
